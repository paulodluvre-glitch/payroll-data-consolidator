import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

st.set_page_config(page_title="Consolidador de Folha de Pagamento", layout="wide")

st.title("📑 Consolidar Folhas de Pagamento")
st.markdown("""
Arraste os arquivos de folha (.xlsx) para consolidar os proventos e descontos por setor e empresa.
""")

def processar_folhas(arquivos_carregados):
    dados_consolidados = []
    
    for arquivo in arquivos_carregados:
        df = pd.read_excel(arquivo, usecols=[1, 4, 18, 20, 22, 25], 
                           names=['EMPRESA', 'COMPETENCIA', 'SETOR', 'TIPO', 'RUBRICA', 'VALOR'])
        
        df = df[df['TIPO'].isin(['P', 'D'])].copy()
        
        df = df[~df['RUBRICA'].astype(str).str.contains('DEPENDENTE.*IRRF.*MENSAL', case=False, regex=True)]
        
        df['SETOR'] = df['SETOR'].apply(lambda x: str(x).split('-')[-1].strip().upper() if pd.notnull(x) else x)
        
        df['COMPETENCIA'] = pd.to_datetime(df['COMPETENCIA'], format='mixed', dayfirst=True).dt.strftime('%d/%m/%Y')
        
        df['VALOR'] = np.where(df['TIPO'] == 'D', -df['VALOR'].astype(float), df['VALOR'].astype(float))
        
        dados_consolidados.append(df)
        
    df_completo = pd.concat(dados_consolidados, ignore_index=True)
    
    proventos_cols = sorted(df_completo[df_completo['TIPO'] == 'P']['RUBRICA'].dropna().unique().tolist())
    descontos_cols = sorted(df_completo[df_completo['TIPO'] == 'D']['RUBRICA'].dropna().unique().tolist())
    
    df_pivot = df_completo.pivot_table(
        index=['EMPRESA', 'SETOR', 'COMPETENCIA'],
        columns='RUBRICA',
        values='VALOR',
        aggfunc='sum'
    ).reset_index()
    
    colunas_ordem = ['EMPRESA', 'SETOR', 'COMPETENCIA'] + proventos_cols + descontos_cols
    df_pivot = df_pivot.reindex(columns=colunas_ordem)
    df_pivot.fillna(0, inplace=True)
    
    df_pivot.insert(3, 'SALDO LIQUIDO', df_pivot[proventos_cols + descontos_cols].sum(axis=1))
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_pivot.to_excel(writer, index=False, startrow=1, sheet_name='Consolidado')
        
        wb = writer.book
        ws = wb['Consolidado']
        
        fill_provento = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fill_desconto = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        fonte_negrito = Font(bold=True)
        alinhamento_centro = Alignment(horizontal="center", vertical="center")
        
        letra_inicio_prov = 5 
        letra_fim_prov = letra_inicio_prov + len(proventos_cols) - 1
        letra_inicio_desc = letra_fim_prov + 1
        letra_fim_desc = letra_inicio_desc + len(descontos_cols) - 1
        
        if len(proventos_cols) > 0:
            celula_prov = ws.cell(row=1, column=letra_inicio_prov, value="PROVENTOS")
            celula_prov.font = fonte_negrito
            celula_prov.fill = fill_provento
            celula_prov.alignment = alinhamento_centro
            ws.merge_cells(start_row=1, start_column=letra_inicio_prov, end_row=1, end_column=letra_fim_prov)
            
        if len(descontos_cols) > 0:
            celula_desc = ws.cell(row=1, column=letra_inicio_desc, value="DESCONTOS")
            celula_desc.font = fonte_negrito
            celula_desc.fill = fill_desconto
            celula_desc.alignment = alinhamento_centro
            ws.merge_cells(start_row=1, start_column=letra_inicio_desc, end_row=1, end_column=letra_fim_desc)
        
        for col in range(1, ws.max_column + 1):
            celula = ws.cell(row=2, column=col)
            celula.font = fonte_negrito
            celula.value = str(celula.value).upper()

    return df_pivot, output.getvalue()

arquivos = st.file_uploader("Suba os arquivos de Folha de Pagamento", type=['xlsx'], accept_multiple_files=True)

if arquivos:
    if st.button("🚀 Gerar Relatório Consolidado"):
        with st.spinner("Processando..."):
            df_final, excel_binario = processar_folhas(arquivos)
            
            st.success("Relatório gerado com sucesso!")
            
            st.dataframe(df_final.replace(0, ''), use_container_width=True)
            
            st.download_button(
                label="📥 Baixar Relatório Consolidado (Excel)",
                data=excel_binario,
                file_name="RELATORIO_CONSOLIDADO_MES.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )