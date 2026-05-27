import streamlit as st
import pandas as pd
import numpy as np
import importlib.util
from io import BytesIO
from pathlib import Path
from tempfile import NamedTemporaryFile
from openpyxl.styles import Font, PatternFill, Alignment

st.set_page_config(page_title="Consolidador de Folha de Pagamento", layout="wide")

st.title("📑 Consolidar Folhas de Pagamento")
st.markdown("""
Arraste os arquivos de folha (.xlsx ou .xls) para consolidar os proventos e descontos por setor e por colaborador.
""")

COLUNAS_PROCESSADAS = [
    'EMPRESA',
    'COMPETENCIA',
    'SETOR',
    'COLABORADOR_ID',
    'COLABORADOR',
    'TIPO',
    'RUBRICA',
    'VALOR',
]
COLUNAS_PADRAO_ENTRADA = ['EMPRESA', 'COMPETENCIA', 'SETOR', 'TIPO', 'RUBRICA', 'VALOR']
COLUNAS_LAYOUT_EVENTOS_SEPARADOS = {
    'nome_emp',
    'cp_competencia',
    'cp_desr_sep',
    'cp_codi_epr',
    'cp_nome_epr',
    'cp_nome_eve_p',
    'cp_eve_val_p',
    'cp_nome_eve_d',
    'cp_eve_val_d',
}


def obter_nome_arquivo(arquivo):
    return getattr(arquivo, "name", str(arquivo))


def obter_extensao_arquivo(arquivo):
    return Path(obter_nome_arquivo(arquivo)).suffix.lower()


def salvar_upload_temporario(arquivo, suffix):
    temp = NamedTemporaryFile(delete=False, suffix=suffix)
    try:
        if hasattr(arquivo, "getvalue"):
            conteudo = arquivo.getvalue()
        else:
            try:
                arquivo.seek(0)
            except Exception:
                pass
            conteudo = arquivo.read()

        temp.write(conteudo)
    finally:
        temp.close()
        try:
            arquivo.seek(0)
        except Exception:
            pass

    return temp.name


def converter_xls_para_xlsx_com_excel(origem):
    import pythoncom
    import win32com.client as win32

    pythoncom.CoInitialize()
    excel = None
    workbook = None
    destino = NamedTemporaryFile(delete=False, suffix=".xlsx")
    destino.close()

    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(
            str(origem),
            UpdateLinks=0,
            ReadOnly=True,
            IgnoreReadOnlyRecommended=True,
            Notify=False,
            CorruptLoad=1,
        )
        workbook.SaveAs(str(destino.name), FileFormat=51)
        workbook.Close(False)
        workbook = None
        return destino.name
    except Exception as erro:
        raise RuntimeError(
            "Nao foi possivel converter o arquivo .xls automaticamente. "
            "Verifique se o Microsoft Excel esta instalado e se o arquivo abre normalmente."
        ) from erro
    finally:
        try:
            if workbook is not None:
                workbook.Close(False)
        except Exception:
            pass
        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass
        pythoncom.CoUninitialize()


def carregar_dataframe_excel(arquivo):
    extensao = obter_extensao_arquivo(arquivo)

    if extensao == '.xls':
        origem_temporaria = None
        convertido = None
        origem_leitura = arquivo
        erros = []

        try:
            if not isinstance(arquivo, (str, Path)):
                origem_temporaria = salvar_upload_temporario(arquivo, ".xls")
                origem_leitura = origem_temporaria

            motores = []
            if importlib.util.find_spec("python_calamine"):
                motores.append("calamine")
            if importlib.util.find_spec("xlrd"):
                motores.append("xlrd")

            for motor in motores:
                try:
                    return pd.read_excel(origem_leitura, engine=motor)
                except Exception as erro:
                    erros.append(f"{motor}: {erro}")

            try:
                convertido = converter_xls_para_xlsx_com_excel(origem_leitura)
                return pd.read_excel(convertido, engine='openpyxl')
            except Exception as erro:
                erros.append(f"excel_com: {erro}")

            raise RuntimeError(
                "Nao foi possivel ler o arquivo .xls automaticamente. "
                "Tente instalar a dependencia python-calamine ou verifique se o Microsoft Excel esta instalado. "
                f"Detalhes: {' | '.join(erros)}"
            )
        finally:
            for caminho in [convertido, origem_temporaria]:
                if caminho and Path(caminho).exists():
                    Path(caminho).unlink(missing_ok=True)

    if extensao == '.xlsx':
        return pd.read_excel(arquivo, engine='openpyxl')

    return pd.read_excel(arquivo)


def normalizar_texto(serie):
    texto = serie.astype("string").str.replace(r"\s+", " ", regex=True).str.strip()
    return texto.replace({"": pd.NA, "nan": pd.NA, "None": pd.NA})


def normalizar_setor(serie):
    texto = normalizar_texto(serie)
    return texto.apply(lambda x: x.split('-')[-1].strip().upper() if pd.notna(x) else x)


def normalizar_competencia(serie):
    texto_original = normalizar_texto(serie)
    datas = pd.to_datetime(serie, errors='coerce', dayfirst=True)

    competencia = texto_original.copy()
    competencia.loc[datas.notna()] = datas.loc[datas.notna()].dt.strftime('%d/%m/%Y')
    return competencia


def normalizar_valor(serie):
    if pd.api.types.is_numeric_dtype(serie):
        return pd.to_numeric(serie, errors='coerce')

    texto = (
        serie.astype("string")
        .str.replace(".", "", regex=False)
        .str.replace(",", ".", regex=False)
        .str.strip()
    )
    texto = texto.replace({"": pd.NA, "nan": pd.NA, "None": pd.NA})
    return pd.to_numeric(texto, errors='coerce')


def obter_coluna_setor_origem(df):
    for coluna in ['cp_desr_sep', 'cp_desr_sep_comp', 'cp_desr_sep_2', 'cp_desr_sep_3', 'cp_depto']:
        if coluna in df.columns and df[coluna].notna().any():
            return coluna
    raise ValueError("O arquivo nao possui uma coluna de setor preenchida.")


def padronizar_dataframe(df):
    df = df.copy()
    if 'COLABORADOR_ID' not in df.columns:
        df['COLABORADOR_ID'] = pd.NA
    if 'COLABORADOR' not in df.columns:
        df['COLABORADOR'] = pd.NA

    df['EMPRESA'] = normalizar_texto(df['EMPRESA'])
    df['COMPETENCIA'] = normalizar_competencia(df['COMPETENCIA'])
    df['SETOR'] = normalizar_setor(df['SETOR'])
    df['COLABORADOR'] = normalizar_texto(df['COLABORADOR']).str.upper()
    df['TIPO'] = normalizar_texto(df['TIPO']).str.upper()
    df['RUBRICA'] = normalizar_texto(df['RUBRICA']).str.upper()
    df['VALOR'] = normalizar_valor(df['VALOR'])
    df['COLABORADOR_ID'] = normalizar_texto(df['COLABORADOR_ID'])

    df = df[df['TIPO'].isin(['P', 'D'])].copy()
    df = df.dropna(subset=['EMPRESA', 'COMPETENCIA', 'SETOR', 'RUBRICA', 'VALOR'])
    df = df[~df['RUBRICA'].str.contains('DEPENDENTE.*IRRF.*MENSAL', case=False, regex=True, na=False)]
    df['VALOR'] = np.where(df['TIPO'] == 'D', -df['VALOR'], df['VALOR'])

    return df[COLUNAS_PROCESSADAS]


def extrair_eventos_de_colunas_separadas(df):
    setor_coluna = obter_coluna_setor_origem(df)
    base_colunas = ['nome_emp', 'cp_competencia', setor_coluna, 'cp_codi_epr', 'cp_nome_epr']
    eventos = []

    for tipo, rubrica_coluna, valor_coluna in [
        ('P', 'cp_nome_eve_p', 'cp_eve_val_p'),
        ('D', 'cp_nome_eve_d', 'cp_eve_val_d'),
    ]:
        if rubrica_coluna not in df.columns or valor_coluna not in df.columns:
            continue

        evento = df[base_colunas + [rubrica_coluna, valor_coluna]].copy()
        evento.columns = ['EMPRESA', 'COMPETENCIA', 'SETOR', 'COLABORADOR_ID', 'COLABORADOR', 'RUBRICA', 'VALOR']
        evento['TIPO'] = tipo
        eventos.append(evento[COLUNAS_PROCESSADAS])

    if not eventos:
        raise ValueError("O arquivo nao possui colunas de proventos/descontos reconhecidas.")

    return pd.concat(eventos, ignore_index=True)


def ler_arquivo_folha(arquivo):
    df = carregar_dataframe_excel(arquivo)
    colunas = {str(col).strip() for col in df.columns}

    if COLUNAS_LAYOUT_EVENTOS_SEPARADOS.issubset(colunas):
        return padronizar_dataframe(extrair_eventos_de_colunas_separadas(df))

    if set(COLUNAS_PROCESSADAS).issubset(colunas):
        return padronizar_dataframe(df[COLUNAS_PROCESSADAS])

    if set(COLUNAS_PADRAO_ENTRADA).issubset(colunas):
        base = df[COLUNAS_PADRAO_ENTRADA].copy()
        if 'COLABORADOR' in df.columns:
            base['COLABORADOR'] = df['COLABORADOR']
        if 'COLABORADOR_ID' in df.columns:
            base['COLABORADOR_ID'] = df['COLABORADOR_ID']
        return padronizar_dataframe(base)

    raise ValueError(
        "Estrutura de planilha nao reconhecida. O arquivo precisa conter as colunas do exportador de folha "
        "ou as colunas padrao EMPRESA, COMPETENCIA, SETOR, TIPO, RUBRICA e VALOR."
    )


def consolidar_eventos(arquivos_carregados):
    dados_consolidados = []
    
    for arquivo in arquivos_carregados:
        df = ler_arquivo_folha(arquivo)
        dados_consolidados.append(df)

    if not dados_consolidados:
        raise ValueError("Nenhum arquivo valido foi enviado para processamento.")

    df_completo = pd.concat(dados_consolidados, ignore_index=True)
    return df_completo.dropna(subset=['RUBRICA'])


def gerar_relatorio(df_completo, colunas_indice, colunas_exibicao):
    proventos_cols = sorted(df_completo[df_completo['TIPO'] == 'P']['RUBRICA'].dropna().unique().tolist())
    descontos_cols = sorted(df_completo[df_completo['TIPO'] == 'D']['RUBRICA'].dropna().unique().tolist())
    df_pivot = df_completo.pivot_table(
        index=colunas_indice,
        columns='RUBRICA',
        values='VALOR',
        aggfunc='sum'
    ).reset_index()

    colunas_para_remover = [col for col in colunas_indice if col not in colunas_exibicao]
    if colunas_para_remover:
        df_pivot = df_pivot.drop(columns=colunas_para_remover)

    colunas_ordem = colunas_exibicao + proventos_cols + descontos_cols
    df_pivot = df_pivot.reindex(columns=colunas_ordem)
    df_pivot.fillna(0, inplace=True)

    df_pivot.insert(len(colunas_exibicao), 'SALDO LIQUIDO', df_pivot[proventos_cols + descontos_cols].sum(axis=1))
    df_pivot.columns.name = None
    colunas_numericas = df_pivot.select_dtypes(include=['number']).columns
    df_pivot[colunas_numericas] = df_pivot[colunas_numericas].round(2)
    return df_pivot, proventos_cols, descontos_cols


def gerar_excel_relatorio(df_pivot, sheet_name, quantidade_colunas_base, proventos_cols, descontos_cols):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_pivot.to_excel(writer, index=False, startrow=1, sheet_name=sheet_name)

        wb = writer.book
        ws = wb[sheet_name]

        fill_provento = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fill_desconto = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        fonte_negrito = Font(bold=True)
        alinhamento_centro = Alignment(horizontal="center", vertical="center")

        letra_inicio_prov = quantidade_colunas_base + 2
        letra_fim_prov = letra_inicio_prov + len(proventos_cols) - 1
        letra_inicio_desc = letra_fim_prov + 1
        letra_fim_desc = letra_inicio_desc + len(descontos_cols) - 1

        if len(proventos_cols) > 0:
            celula_prov = ws.cell(row=1, column=letra_inicio_prov, value="PROVENTOS")
            celula_prov.font = fonte_negrito
            celula_prov.fill = fill_provento
            celula_prov.alignment = alinhamento_centro
            ws.merge_cells(start_row=1, start_column=letra_inicio_prov, end_row=1, end_column=letra_fim_prov)

        if len(descontos_cols) > 0:
            celula_desc = ws.cell(row=1, column=letra_inicio_desc, value="DESCONTOS")
            celula_desc.font = fonte_negrito
            celula_desc.fill = fill_desconto
            celula_desc.alignment = alinhamento_centro
            ws.merge_cells(start_row=1, start_column=letra_inicio_desc, end_row=1, end_column=letra_fim_desc)

        for col in range(1, ws.max_column + 1):
            celula = ws.cell(row=2, column=col)
            celula.font = fonte_negrito
            celula.value = str(celula.value).upper()

    return output.getvalue()


def processar_folhas(arquivos_carregados):
    df_completo = consolidar_eventos(arquivos_carregados)
    df_setor, proventos_cols, descontos_cols = gerar_relatorio(
        df_completo,
        ['EMPRESA', 'SETOR', 'COMPETENCIA'],
        ['EMPRESA', 'SETOR', 'COMPETENCIA'],
    )

    df_com_colaborador = df_completo.dropna(subset=['COLABORADOR', 'COLABORADOR_ID']).copy()
    if df_com_colaborador.empty:
        raise ValueError("Os arquivos enviados nao possuem dados suficientes para montar o relatorio por colaborador.")

    df_colaborador, proventos_cols_colab, descontos_cols_colab = gerar_relatorio(
        df_com_colaborador,
        ['EMPRESA', 'COLABORADOR_ID', 'COLABORADOR', 'SETOR', 'COMPETENCIA'],
        ['EMPRESA', 'COLABORADOR', 'SETOR', 'COMPETENCIA'],
    )

    excel_setor = gerar_excel_relatorio(
        df_setor,
        'Consolidado_Setor',
        3,
        proventos_cols,
        descontos_cols,
    )
    excel_colaborador = gerar_excel_relatorio(
        df_colaborador,
        'Consolidado_Colaborador',
        4,
        proventos_cols_colab,
        descontos_cols_colab,
    )

    return {
        'setor': {'df': df_setor, 'excel': excel_setor},
        'colaborador': {'df': df_colaborador, 'excel': excel_colaborador},
    }

arquivos = st.file_uploader("Suba os arquivos de Folha de Pagamento", type=['xlsx', 'xls'], accept_multiple_files=True)
aba_setor, aba_colaborador = st.tabs(["Por Setor", "Por Colaborador"])

if arquivos:
    if st.button("🚀 Gerar Relatório Consolidado"):
        with st.spinner("Processando..."):
            try:
                relatorios = processar_folhas(arquivos)
            except Exception as erro:
                st.error(f"Falha ao processar a planilha: {erro}")
            else:
                st.success("Relatório gerado com sucesso!")

                with aba_setor:
                    st.dataframe(relatorios['setor']['df'].replace(0, ''), use_container_width=True)
                    st.download_button(
                        label="📥 Baixar Relatório por Setor (Excel)",
                        data=relatorios['setor']['excel'],
                        file_name="RELATORIO_CONSOLIDADO_SETOR.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

                with aba_colaborador:
                    st.dataframe(relatorios['colaborador']['df'].replace(0, ''), use_container_width=True)
                    st.download_button(
                        label="📥 Baixar Relatório por Colaborador (Excel)",
                        data=relatorios['colaborador']['excel'],
                        file_name="RELATORIO_CONSOLIDADO_COLABORADOR.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
